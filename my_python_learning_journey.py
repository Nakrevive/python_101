#My learning journe with python


import string


#formatted strings

first = 'Arnaud'
last = 'Ntwari'
message = first + '[' + last + ']' + ' is a dancer'
msg = f'{first}[ {last}] is a dancer'

print(msg.upper())#Output: Arnaud[ Ntwari] is a dancer


course = 'Python for Beginners'
another = course[:]
print(another)



name = 'Jennifer'
print(name[1:-1]) ; # Output: enner

# string methods

course = 'Python for Beginners'
 
print(course.upper())
print(course.lower())
print(course)

print(course.find('P')) #Output: 0
print(course.replace('Beginners', 'Absolute Beginners'))
print('Python' in course)

#operator precedence
x = 10 + 3 * 2**2 
print(x) #Output: 16

# math functions
x = 2.9
print(round(x)) #Outpu

x= 2.9
print(abs(-2.9))#Output: 2.9
# import math
import math
print(math.ceil(2.9)) #Output: 3
print(math.floor(2.9)) #Output: 2

# if statements
is_hot = False
is_cold = False
if is_hot:
    print("It's a hot day")
    print("Drink plenty of water")
elif is_cold:
    print("It's a cold day")
    print("Wear warm clothes")
    
else:
    
    print("It's a cold day")
    print("Wear warm clothes") 
print("Enjoy your day") 

#exercise
price = 1000000
has_good_credit = True
if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f'Down payment: ${down_payment}')

#logical operators
has_high_income = True
has_criminal_record = False
if has_high_income and not has_criminal_record:
        print("Eligible for loan")
        
 
 #comparison operators
temperature = 35
if temperature > 30:
    print("It's a hot day")       
else:
    print("It's not a hot day")
        
        #exercise
name = "ntwarii"
if len(name) < 3:
    print("Name must be at least 3 characters")
elif len(name) > 50:
    print("Name can be a maximum of 50 characters")
else:
    print("Name looks good!")
    
    #weight converter
    weight = int(input("Weight: "))
    unit = input("(L)bs or (K)g: ")
    if unit.upper() == "L":
        converted = weight * 0.45
        print(f"You are {converted} kilos")
    else:
        converted = weight / 0.45
        print(f"You are {converted} pounds");
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        
        # while loops
i = 1
while i <= 5:
    print('*' *i)
    i = i + 1
print("Done")
        
    




    
        # guessing game
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < 3:
    guess =int(input("Guess: "))
    guess_count += 1
    if guess == secret_number:
        print("You won!")
        break
    else:
        print("Sorry, you failed!")
    
    
    
    
    
    #CAR GAME
            
        command = ""
staeted = False
while command != "quit":
    command = input(">").lower()
   if command == "start":
       if staeted:
           print("Car is already started!")
        print("Car started...Ready to go!")
    elif command == "stop":
        if not started
        
        print("Car is already stopped!")
        else:
started = False
print("Car stopped.")
elif command == "help":
print("""start - to start the car
stop - to stop the car
quit - to exit the game""")
elif command == "quit":
    break
    else:
print("I don't understand that...")





#for loops
for item in 'python' :
    print(item);
    
 for item in range (5,10,2) :
    print(item) 
    
    #exercise
    prices = [10, 20, 30]

total = 0
for price in prices:
    total +=  price
    print(f"Total: {total}") 
    
     #nested loops
for x in range(4):
    for y in range  (3):
        print(f'({x}, {y})')
        
        
        #exercise
numbers = [5,2,5,2,2]
for x_count in numbers:
    output = ''
    for count in range (x_count):
        output += 'x'
    print(output)
    
    
    #lists
    names = ['ntwari', 'bisizi', 'karire', 'ganza', 'ngabo']
names[2] = 'karl'
print(names)

# exercise
numbers = [3,5,4,5,8,10]
max = numbers[0]
for number in numbers:
    if number > max :
        max = number
        print(max)
        
#2D lists
atrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
matrix[0] [1] = 30
print(matrix[0] [1])

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]


for row in matrix:
    for item in row:
        print (item)
        
        #list methods
        
        
        numbers = [5, 2, 1, 7, 4]
numbers.append(20)
print(numbers)

numbers = [5, 2, 1, 7, 4]
numbers.insert(0,10)
print(numbers)


numbers = [5, 2, 1, 7, 4]
numbers2 = numbers.copy()
numbers.append(20)
print(numbers) 

#exercise
numbers = [2,2,4,6,3,4,6,1]
uniques = []
for number in numbers:
    if number not in uniques:
        uniques.append(number)
        
print(uniques)


#tuples
numbers = (1, 2, 3)
print(numbers[0])


numbers = (1, 2, 3)
numbers[0] = 10
print(numbers[0])
## erro

#unpacking
coordinates = (1, 2, 3)
x,y,z = coordinates
print(y)

#dictionaries
customer = {
    "name": "ntwari arnaud",
    "age": 23,
  "is_verified": True
}
customer["salary"] = 5000
print(customer["salary"])

#exercise
phone =input("phone = ")
number_in_words = {"0": "zero",
                   "1": "one",
                   "2": "two",
                   "3": "three",
                   "4": "four",
                     "5": "five",
                     "6": "six",
                     "7": "seven",
                     "8": "eight", 
                     "9": "nine"}
output = ""

for ch in phone:
    output += number_in_words.get(ch, "!") + " "
print(output)

#emoji converter
message = input(">")
words = message.split(' ')
emojis = {":)": "😂",
          ":(": "😒"}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
    print(output)
    
    # functions
    
def greet_user():
    print('Hi there')
    print('Welcome aboard')
    
    
print("start")
greet_user()
print("Finish")

#parameters

def greet_user(name):
    print(f'Hi {name} there')
    print('Welcome aboard')
    
    
print("start")
greet_user("ntwari")
greet_user("karire")
print("Finish")

#keyword arguments
def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name} there')
    print('Welcome aboard')
    
    
print("start")
greet_user("ntwari", "arnaud")


print("Finish")

#return statements


def square(number):
    return number * number

print(square(3))
  
  # creating a reusable function
    def emoji_converter(message):
    words = message.split(" ")
    emojis = {
        ":)": "😂",
        ":(": "😒"
    }
    output = ""
    for word in words:
      output += emojis.get(word,word) + ""
    return output


    message = input(">")
    print(emoji_converter(message)) 
    
    #execption
try:
    age = int(input('age: '))
    income = 30000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print('agae cannot be o.')
except ValueError:
    print('invalid value')
    
    #comments
#comments are goood but should not be too much and 
# obvious some are obvious and no
# need to write them
print("sky is blue")


#classes
class point:
    def move(self):
        print("move")
    def draw(self):
        print("draw")
        
        
point1 = point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()

point2 = point()
point2.x = 1
print(point2.x)
    
    
    
    #constructors
class point:
    def __init__(self,x,y):
        self.x = x
        self.y = y
    def move(self):
        print("move")
    def draw(self):
        print("draw")
        
        
point = point(10,20)
point.x = 11
print(point.x) 

#exercise       
class Person:
    def __init__(self,name):
        self.name = name
    def talk(self):
        print(f"Hi, I am {self.name}")


ntwari = Person("Ntwari Arnaud")
print(ntwari.name)
ntwari.talk()

rwizihirangabo = Person("Rwizihirangabo")
rwizihirangabo.talk()



# inheritance mechanism for reusing code
class Mammal:
    def walk(self):
        print("walk")
    
class Dog(Mammal):
    def bark(self):
        print("bark")

        
class cat(Mammal):
    def be_annoying(self):
        print("annoying")

dog1 = Dog()
dog1.walk()

cat1 = cat()
cat1.be_annoying()




    #converters
import converter
from converters import kg_to_lbs
print(converters.kg_to_lbs(100))

#exercise
import utils
from utils import find_max
numbers = [10,3,6,2]
# max = find_max(numbers)
print(max)




#packages
#two ways of importing
import ecommerce.shipping
ecommerce.shipping.calc_shipping()


#2nd approach
from ecommerce.shipping import calc_shipping
calc_shipping

#or
 from ecommerce import shipping
 shipping.calc_shipping()


 #Generating random values
 import random 

 for i in range(3):
    print(random.randint(10,20))
    
    
    
    #2
    # Generating random values
import random

members = ['Ntwari', 'Nora','Norine','Ighor']
leader = random.choice(members)
print(leader)

# exercise
import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first,second
   
   
dice = Dice()
print(dice.roll())  


#files and directories
from pathlib import Path
path = Path()
for file in path.glob('*'):
    print(file)
    

#Pypi and Pip
  #instslling openpyxl
  
  
  
  
  #projects
  
  #1 excel spreadsheets
import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']


    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value* 0.9
        corrected_price_cell = sheet.cell(row, 4) 
        corrected_price_cell.value = corrected_price
    values = Reference(sheet,min_row=2,
            max_row=sheet.max_row,
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename) 
  
  #machine learning
 # in jupyter
import pandas as pd
from sklearn.tree import DecisionTreeClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score

music_data = pd.read_csv('C:/Users/educa/Desktop/music.csv')
x = music_data.drop(columns=['genre'])
y = music_data['genre']
x_train, x_test, y_train, y_test = train_test_split(x,y,test_size = 0.2)

model = DecisionTreeClassifier()
model.fit(x_train, y_train)
predictions = model.predict(x_test)

score = accuracy_score(y_test, predictions)
score
  
  
  
    
    
    
    
